#!/usr/bin/env python3
"""
German Tax Invoice Extraction System

Extracts tax-relevant information from German vendor PDF invoices
and maintains an Excel tracking sheet.
"""

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Constants
NEW_FOLDER = Path("./new")
ARCHIVE_FOLDER = Path("./archive")
EXCEL_FILE = Path("./tax_records.xlsx")
LOG_FILE = Path("./extraction_log.txt")

# German expense categories
EXPENSE_CATEGORIES = [
    "Büromaterial",
    "Software",
    "Reisekosten",
    "Marketing",
    "Telefon/Internet",
    "Miete",
    "Versicherung",
    "Weiterbildung",
    "Beratung",
    "Sonstiges"
]

# Color definitions
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Regex patterns for German invoices
PATTERNS = {
    # Currency patterns: "1.234,56 €", "€ 1.234,56", "1234.56", "1.234,56"
    'currency': re.compile(r'€?\s*(\d{1,3}(?:[.\s]\d{3})*,\d{2}|\d+\.\d{2})\s*€?'),

    # Date patterns: DD.MM.YYYY, DD.MM.YY
    'date': re.compile(r'\b(\d{1,2}\.(?:0[1-9]|1[0-2])\.(?:\d{4}|\d{2}))\b'),

    # Invoice number patterns
    'invoice_number': re.compile(
        r'(?:Rechnung|Rechnungs?[-\s]?Nr\.?|Invoice)[\s:]*([A-Z0-9\-/]+)',
        re.IGNORECASE
    ),

    # VAT patterns (German)
    'vat_rate': re.compile(r'(\d{1,2})[%\s]*(?:MwSt|Mehrwertsteuer|USt|Umsatzsteuer)', re.IGNORECASE),

    # German company suffixes
    'company_suffix': re.compile(r'\b(GmbH|AG|UG|KG|OHG|e\.V\.)\b'),
}


def log_message(message: str, level: str = "INFO"):
    """Log message to file with timestamp."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {level}: {message}\n"

    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(log_entry)

    if level == "ERROR":
        print(f"❌ {message}")
    elif level == "WARNING":
        print(f"⚠️  {message}")
    else:
        print(f"✓ {message}")


def sanitize_vendor_name(vendor_name: str) -> str:
    """
    Sanitize vendor name for use as folder name.
    - Convert to lowercase
    - Replace umlauts: ä→ae, ö→oe, ü→ue, ß→ss
    - Replace spaces and special chars with underscores
    """
    if not vendor_name:
        return "unknown_vendor"

    # Replace umlauts
    replacements = {
        'ä': 'ae', 'Ä': 'ae',
        'ö': 'oe', 'Ö': 'oe',
        'ü': 'ue', 'Ü': 'ue',
        'ß': 'ss'
    }

    for old, new in replacements.items():
        vendor_name = vendor_name.replace(old, new)

    # Convert to lowercase
    vendor_name = vendor_name.lower()

    # Replace spaces and special characters with underscores
    vendor_name = re.sub(r'[^\w\s-]', '', vendor_name)
    vendor_name = re.sub(r'[\s-]+', '_', vendor_name)

    return vendor_name.strip('_')


def parse_german_currency(value: str) -> Optional[float]:
    """
    Parse German currency format to float.
    Handles: "1.234,56", "1234,56", "1234.56"
    """
    if not value:
        return None

    # Remove currency symbol and spaces
    value = value.replace('€', '').strip()

    # Check if it's German format (comma as decimal separator)
    if ',' in value:
        # Remove thousand separators (. or space)
        value = value.replace('.', '').replace(' ', '')
        # Replace comma with dot
        value = value.replace(',', '.')
    else:
        # Might be US format or integer, remove spaces
        value = value.replace(' ', '')

    try:
        return float(value)
    except ValueError:
        return None


def parse_german_date(date_str: str) -> Optional[str]:
    """
    Parse German date format to YYYY-MM-DD.
    Handles: DD.MM.YYYY, DD.MM.YY
    """
    if not date_str:
        return None

    try:
        # Try DD.MM.YYYY
        if len(date_str.split('.')[-1]) == 4:
            dt = datetime.strptime(date_str, "%d.%m.%Y")
        else:
            # Try DD.MM.YY
            dt = datetime.strptime(date_str, "%d.%m.%y")

        return dt.strftime("%d.%m.%Y")
    except ValueError:
        return None


def extract_vendor_name(text: str) -> Optional[str]:
    """
    Extract vendor name from invoice text.
    Look for company suffixes (GmbH, AG, etc.) and nearby text.
    """
    lines = text.split('\n')[:15]  # Check first 15 lines

    # Look for lines with company suffixes
    for line in lines:
        if PATTERNS['company_suffix'].search(line):
            # Clean up the line
            line = line.strip()
            # Remove common prefixes
            line = re.sub(r'^(Von:|From:|Lieferant:|Aussteller:)\s*', '', line, flags=re.IGNORECASE)

            if len(line) > 3 and len(line) < 100:
                return line

    # Fallback: look for capitalized words in first few lines
    for line in lines[:5]:
        line = line.strip()
        # Skip very short or very long lines
        if 5 < len(line) < 80:
            # Check if line has multiple capitalized words
            words = line.split()
            if len(words) >= 2 and sum(1 for w in words if w[0].isupper()) >= 2:
                return line

    return None


def extract_invoice_data(pdf_path: Path) -> Dict:
    """
    Extract tax-relevant data from PDF invoice.
    Returns dict with extracted fields and extraction status.
    """
    data = {
        'filename': pdf_path.name,
        'date': None,
        'vendor': None,
        'invoice_number': None,
        'net': None,
        'vat_rate': None,
        'vat_amount': None,
        'gross': None,
        'category': '',
        'extraction_status': 'OK',
        'notes': ''
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from first few pages
            text = ""
            for page in pdf.pages[:3]:  # Check first 3 pages
                text += page.extract_text() or ""

            if not text.strip():
                data['extraction_status'] = 'MANUAL_REVIEW_NEEDED'
                data['notes'] = 'No text extracted from PDF'
                return data

            # Extract vendor name
            vendor = extract_vendor_name(text)
            if vendor:
                data['vendor'] = vendor
            else:
                data['vendor'] = 'Unknown Vendor'
                data['extraction_status'] = 'UNCERTAIN'
                data['notes'] += 'Vendor name unclear; '

            # Extract invoice number
            inv_match = PATTERNS['invoice_number'].search(text)
            if inv_match:
                data['invoice_number'] = inv_match.group(1).strip()
            else:
                data['extraction_status'] = 'UNCERTAIN'
                data['notes'] += 'Invoice number not found; '

            # Extract date (look for recent dates)
            dates = PATTERNS['date'].findall(text)
            for date_str in dates:
                parsed_date = parse_german_date(date_str)
                if parsed_date:
                    data['date'] = parsed_date
                    break

            if not data['date']:
                data['extraction_status'] = 'UNCERTAIN'
                data['notes'] += 'Date not found; '

            # Extract VAT rate
            vat_rate_match = PATTERNS['vat_rate'].search(text)
            if vat_rate_match:
                data['vat_rate'] = f"{vat_rate_match.group(1)}%"
            else:
                # Default to 19% for Germany
                data['vat_rate'] = '19%'
                data['notes'] += 'VAT rate assumed 19%; '

            # Extract currency amounts
            amounts = []
            for match in PATTERNS['currency'].finditer(text):
                amount = parse_german_currency(match.group(1))
                if amount and amount > 0:
                    amounts.append(amount)

            # Sort amounts to identify likely net/vat/gross
            amounts = sorted(set(amounts))

            if len(amounts) >= 3:
                # Likely: smallest amounts first, gross is largest
                data['gross'] = amounts[-1]

                # Try to find VAT rate to calculate net
                vat_rate_num = 19  # default
                if data['vat_rate']:
                    vat_rate_num = int(data['vat_rate'].replace('%', ''))

                # Calculate expected net from gross
                expected_net = data['gross'] / (1 + vat_rate_num / 100)

                # Find closest amount to expected net
                for amt in amounts:
                    if abs(amt - expected_net) < expected_net * 0.1:  # Within 10%
                        data['net'] = amt
                        data['vat_amount'] = data['gross'] - data['net']
                        break

                if not data['net']:
                    # Fallback: assume amounts[-2] is net
                    data['net'] = amounts[-2] if len(amounts) >= 2 else None
                    if data['net']:
                        data['vat_amount'] = data['gross'] - data['net']

            elif len(amounts) == 1:
                # Only one amount found, assume it's gross
                data['gross'] = amounts[0]
                vat_rate_num = 19
                if data['vat_rate']:
                    vat_rate_num = int(data['vat_rate'].replace('%', ''))
                data['net'] = data['gross'] / (1 + vat_rate_num / 100)
                data['vat_amount'] = data['gross'] - data['net']
                data['extraction_status'] = 'UNCERTAIN'
                data['notes'] += 'Only one amount found, calculated net/VAT; '

            # Validate amounts
            if data['net'] and data['vat_amount'] and data['gross']:
                calculated_gross = data['net'] + data['vat_amount']
                if abs(calculated_gross - data['gross']) > 0.01:
                    data['extraction_status'] = 'UNCERTAIN'
                    data['notes'] += 'Amount validation failed; '
            else:
                data['extraction_status'] = 'MANUAL_REVIEW_NEEDED'
                data['notes'] += 'Missing amount data; '

    except Exception as e:
        data['extraction_status'] = 'MANUAL_REVIEW_NEEDED'
        data['notes'] = f'Extraction error: {str(e)}'

    return data


def setup_excel_file():
    """Create Excel file with proper structure if it doesn't exist."""
    if EXCEL_FILE.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Headers
    headers = [
        "Filename", "Date", "Vendor", "Invoice_Number", "Net",
        "VAT_Rate", "VAT_Amount", "Gross", "Category",
        "Extraction_Status", "Notes"
    ]
    ws.append(headers)

    # Format header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(1, col_num)
        cell.font = Font(bold=True)

    # Set column widths
    column_widths = {
        'A': 30,  # Filename
        'B': 12,  # Date
        'C': 25,  # Vendor
        'D': 15,  # Invoice_Number
        'E': 12,  # Net
        'F': 10,  # VAT_Rate
        'G': 12,  # VAT_Amount
        'H': 12,  # Gross
        'I': 15,  # Category
        'J': 18,  # Extraction_Status
        'K': 40,  # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add data validation for Category column
    dv = DataValidation(type="list", formula1=f'"{",".join(EXPENSE_CATEGORIES)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'I2:I1000')  # Apply to first 1000 rows

    wb.save(EXCEL_FILE)
    log_message(f"Created Excel file: {EXCEL_FILE}")


def get_existing_filenames() -> set:
    """Get set of filenames already processed in Excel."""
    if not EXCEL_FILE.exists():
        return set()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    filenames = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            filenames.add(row[0])

    wb.close()
    return filenames


def append_to_excel(data: Dict):
    """Append invoice data to Excel file."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find next empty row
    next_row = ws.max_row + 1

    # Append data
    row_data = [
        data['filename'],
        data['date'],
        data['vendor'],
        data['invoice_number'],
        data['net'],
        data['vat_rate'],
        data['vat_amount'],
        data['gross'],
        data['category'],
        data['extraction_status'],
        data['notes'].strip()
    ]

    ws.append(row_data)

    # Format currency columns (E, G, H - Net, VAT_Amount, Gross)
    for col in ['E', 'G', 'H']:
        cell = ws[f'{col}{next_row}']
        if cell.value:
            cell.number_format = '#,##0.00 €'

    # Format date column (B)
    cell = ws[f'B{next_row}']
    if cell.value:
        cell.number_format = 'DD.MM.YYYY'

    # Apply color coding based on extraction status
    if data['extraction_status'] == 'UNCERTAIN':
        for col in range(1, 12):
            ws.cell(next_row, col).fill = YELLOW_FILL
    elif data['extraction_status'] == 'MANUAL_REVIEW_NEEDED':
        for col in range(1, 12):
            ws.cell(next_row, col).fill = RED_FILL

    wb.save(EXCEL_FILE)


def move_to_archive(pdf_path: Path, vendor_name: str):
    """Move PDF to vendor-specific archive folder."""
    sanitized_vendor = sanitize_vendor_name(vendor_name)
    vendor_folder = ARCHIVE_FOLDER / sanitized_vendor
    vendor_folder.mkdir(parents=True, exist_ok=True)

    destination = vendor_folder / pdf_path.name

    # Handle duplicate filenames
    if destination.exists():
        base_name = pdf_path.stem
        suffix = pdf_path.suffix
        counter = 1
        while destination.exists():
            destination = vendor_folder / f"{base_name}_{counter}{suffix}"
            counter += 1

    shutil.move(str(pdf_path), str(destination))
    return destination


def process_invoices():
    """Main processing function: extract data from PDFs and update Excel."""
    # Ensure folders exist
    NEW_FOLDER.mkdir(exist_ok=True)
    ARCHIVE_FOLDER.mkdir(exist_ok=True)

    # Setup Excel file
    setup_excel_file()

    # Get list of PDFs to process
    pdf_files = list(NEW_FOLDER.glob("*.pdf"))

    if not pdf_files:
        print("No PDF files found in ./new/ folder.")
        return

    print(f"Found {len(pdf_files)} PDF(s) to process.\n")

    # Get already processed filenames
    existing_filenames = get_existing_filenames()

    processed_count = 0
    skipped_count = 0
    error_count = 0

    for pdf_path in pdf_files:
        # Check if already processed
        if pdf_path.name in existing_filenames:
            print(f"⏭️  Skipping {pdf_path.name} (already processed)")
            skipped_count += 1
            continue

        print(f"Processing: {pdf_path.name}")

        try:
            # Extract data
            data = extract_invoice_data(pdf_path)

            # Append to Excel
            append_to_excel(data)

            # Move to archive
            archive_path = move_to_archive(pdf_path, data['vendor'])

            # Log results
            status_icon = "✓" if data['extraction_status'] == 'OK' else "⚠️" if data['extraction_status'] == 'UNCERTAIN' else "❌"
            print(f"  {status_icon} Vendor: {data['vendor']}")
            print(f"  {status_icon} Gross: {data['gross']:.2f} € " if data['gross'] else "  ⚠️  Gross: N/A")
            print(f"  {status_icon} Status: {data['extraction_status']}")
            print(f"  ✓ Archived to: {archive_path.parent.name}/{archive_path.name}\n")

            log_message(f"Processed {pdf_path.name} -> {archive_path}")
            processed_count += 1

        except Exception as e:
            error_msg = f"Failed to process {pdf_path.name}: {str(e)}"
            log_message(error_msg, "ERROR")
            error_count += 1
            print()

    # Summary
    print("=" * 50)
    print(f"Processing complete!")
    print(f"  Processed: {processed_count}")
    print(f"  Skipped: {skipped_count}")
    print(f"  Errors: {error_count}")
    print(f"  Excel file: {EXCEL_FILE}")
    print("=" * 50)


def show_report():
    """Display summary statistics from Excel file."""
    if not EXCEL_FILE.exists():
        print("No Excel file found. Process some invoices first.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Collect statistics
    total_invoices = 0
    vendor_counts = {}
    month_counts = {}
    total_gross = 0
    status_counts = {'OK': 0, 'UNCERTAIN': 0, 'MANUAL_REVIEW_NEEDED': 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        total_invoices += 1

        # Vendor statistics
        vendor = row[2] or 'Unknown'
        vendor_counts[vendor] = vendor_counts.get(vendor, 0) + 1

        # Month statistics
        if row[1]:  # Date
            try:
                if isinstance(row[1], str):
                    date_obj = datetime.strptime(row[1], "%d.%m.%Y")
                else:
                    date_obj = row[1]
                month_key = date_obj.strftime("%Y-%m")
                month_counts[month_key] = month_counts.get(month_key, 0) + 1
            except:
                pass

        # Gross total
        if row[7]:  # Gross amount
            try:
                total_gross += float(row[7])
            except:
                pass

        # Status counts
        status = row[9] or 'OK'
        if status in status_counts:
            status_counts[status] += 1

    wb.close()

    # Display report
    print("\n" + "=" * 50)
    print("INVOICE TRACKING REPORT")
    print("=" * 50)

    print(f"\nTotal Invoices: {total_invoices}")
    print(f"Total Gross Amount: {total_gross:,.2f} €")

    print(f"\nExtraction Status:")
    for status, count in status_counts.items():
        if count > 0:
            print(f"  {status}: {count}")

    print(f"\nTop Vendors:")
    sorted_vendors = sorted(vendor_counts.items(), key=lambda x: x[1], reverse=True)
    for vendor, count in sorted_vendors[:10]:
        print(f"  {vendor}: {count} invoice(s)")

    print(f"\nInvoices by Month:")
    sorted_months = sorted(month_counts.items(), reverse=True)
    for month, count in sorted_months[:12]:
        print(f"  {month}: {count} invoice(s)")

    print("\n" + "=" * 50)


def list_new_pdfs():
    """List all PDFs in the new/ folder without processing."""
    NEW_FOLDER.mkdir(exist_ok=True)
    pdf_files = list(NEW_FOLDER.glob("*.pdf"))

    if not pdf_files:
        print("No PDF files found in ./new/ folder.")
        return

    print(f"\nFound {len(pdf_files)} PDF(s) in ./new/ folder:\n")
    for pdf_path in pdf_files:
        size_kb = pdf_path.stat().st_size / 1024
        print(f"  • {pdf_path.name} ({size_kb:.1f} KB)")

    print()


def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description="German Tax Invoice Extraction System",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python invoice_extractor.py process    # Process all PDFs in ./new/
  python invoice_extractor.py report     # Show summary statistics
  python invoice_extractor.py list       # List PDFs without processing
        """
    )

    parser.add_argument(
        'command',
        choices=['process', 'report', 'list'],
        help='Command to execute'
    )

    args = parser.parse_args()

    if args.command == 'process':
        process_invoices()
    elif args.command == 'report':
        show_report()
    elif args.command == 'list':
        list_new_pdfs()


if __name__ == "__main__":
    main()
